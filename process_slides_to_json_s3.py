import os
import json
from collections import defaultdict
import openpyxl

slides_content = {1: ['sample_text_02.assets/slide_01_img_1.png', 'sample_text_02.assets/slide_01_img_4.png'],
                  2: ['мем', 'сегодня', ':eye:', 'sample_text_02.assets/slide_02_img_1.png', 'интеллектуально?',
                      ':heart:', ':heart:', ':heart:', 'я', 'не', 'могу', 'не', 'говорить', 'частицы', 'моей',
                      'жизни', 'очень', 'плохо', 'на', 'самом', 'деле', ':sob:', ':sob:', ':sob:', 'геншин', 'или',
                      'бсд', 'я', 'просто', 'ничего', 'не', 'понимаю', ':disappointed_relieved:'],
                  3: ['РПП', 'расстройство', 'пищевого', 'поведения', 'психическое', 'отклонение', 'РПП', 'с', 'РПП',
                      'плачут', 'впадают', 'в', 'истерику', 'не', 'в', 'силах', 'остановится'],
                  4: ['прекрасного', 'счастья', 'любви', 'тепла', 'поддержки!', 'эстетика???',
                      'sample_text_02.assets/slide_04_img_1.png'],
                  7: ['sample_text_02.assets/slide_07_img_1.png'],
                  8: ['клубок', 'из', 'собственных', 'мыслей', 'замок', 'сковывающий', 'Замок', 'открытой',
                      'ладонью', 'ключ', 'близким', 'людям', 'замок'],
                  9: ['надо', 'мной', 'в', 'моей', 'беспорядок', 'как', 'клубок', 'в', 'моей', 'голове', 'метафорой',
                      'что', 'я', 'Мой', 'замок', 'моя', 'замкнутость', 'и', 'недоверие', 'замку', 'ключ'],
                  10: ['клубок', 'мыслей', 'в', 'голове', 'их', 'беспорядочность', 'клубок', 'Замок', 'замкнутость',
                       'и', 'недоверие', 'замку', 'ключ'],
                  11: ['sample_text_02.assets/slide_11_img_1.png', 'sample_text_02.assets/slide_11_img_3.png'],
                  12: ['современных', 'школьников', '—', 'детям', 'ребёнок'],
                  13: ['агрессивна', 'враждебным', 'Подросткам', 'избавиться', 'от', 'оков', 'этого', 'мира', 'к',
                       'мыслям', 'о', 'смерти', 'от', 'настигающих', 'проблем'],
                  14: ['современных', 'детей', 'детей', 'подростков', 'ребятам'],
                  16: ['Подросткам', 'с', 'одиночеством', 'и', 'непониманием', 'мира', 'вокруг', 'дети', 'сенситивны',
                       'поддержка', 'их', 'любовь', 'понимание', 'Дети', 'тепла', 'и', 'доброты', 'В', 'сенсорном',
                       'отношении', 'ребёнок', 'тесные', 'объятия', 'и', 'тактильны', 'контакт'],
                  17: ['Ребенок', 'страдает', 'от', 'этого,', 'ему', 'от', 'этого', 'особенно', 'плохо'],
                  18: ['ПРЛ', 'у', 'детей', 'диагнозов', 'c', 'пограничным', 'расстройством', 'личности', 'детей',
                       '«в', 'больничку»', 'с', 'ПРЛ', 'антидепрессанты', 'препараты', 'помощь', 'и', 'поддержка',
                       'любовь', 'и', 'тепло'],
                  19: ['интеллект', 'усталость', 'жизнь', 'красота', 'тревога', 'непонимание', 'думать', 'любовь',
                       'личность', 'скованность'],
                  20: ['не люблю людей', 'всё', 'страх', 'избавление', 'суицид', 'жизнь', 'смерть', 'чувствовать',
                       'плохо', 'дом', 'человек', 'сильно', 'эстетика', 'одиночество'],
                  21: ['люди', 'один', 'надеяться', 'штиль', 'красота']
                  }


def normalize_path(path):
    if path is None:
        return ""
    if path.startswith("UI/data/emoji/"):
        emoji_name = os.path.splitext(os.path.basename(path))[0]
        return f":{emoji_name}:"
    elif "sample_text_02.assets/" in path:
        return path.split("/")[-1]
    return path


def process_xlsx_files(folder_path, results_folder, slides_content):
    if not os.path.exists(results_folder):
        os.makedirs(results_folder)

    for filename in os.listdir(folder_path):
        if filename.endswith('.xlsx'):
            data_collection = {}
            file_path = os.path.join(folder_path, filename)
            workbook = openpyxl.load_workbook(file_path)

            for sheet_number, words in slides_content.items():
                slide_data = defaultdict(float)
                normalized_words = [normalize_path(word) for word in words]

                if sheet_number <= len(workbook.sheetnames):
                    sheet = workbook.worksheets[sheet_number - 1]

                    for row in sheet.iter_rows(min_row=2, values_only=True):
                        word, delay_time_str = row[0], row[1]
                        if word is None or delay_time_str is None:
                            continue
                        try:
                            delay_time = float(delay_time_str)
                        except ValueError:
                            continue

                        normalized_word = normalize_path(word)
                        if normalized_word in normalized_words:
                            count = sum(1 for k in slide_data if k.startswith(normalized_word))
                            unique_key = f"{normalized_word}_{count + 1}"
                            slide_data[unique_key] = delay_time

                final_data = dict(slide_data)
                data_collection[f'Slide {sheet_number}'] = final_data

            json_file_path = os.path.join(results_folder, f"{os.path.splitext(filename)[0]}.json")
            with open(json_file_path, 'w', encoding='utf-8') as json_file:
                json.dump(data_collection, json_file, ensure_ascii=False, indent=4)


file_type_input = int(input('Введите 1 (группа нормы) или 2 (контрольная группа): '))
if file_type_input == 1:
    folder_path = 'folder_группа_нормы'
    results_folder = 'results_группа_нормы'
elif file_type_input == 2:
    folder_path = 'folder_контрольная_группа'
    results_folder = 'results_контрольная_группа'
else:
    raise ValueError("Неверный ввод")

process_xlsx_files(folder_path, results_folder, slides_content)
