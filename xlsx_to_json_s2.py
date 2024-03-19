import os
import json
from openpyxl import load_workbook


def process_xlsx_files(folder_path):
    json_folder = folder_path + '_json'
    os.makedirs(json_folder, exist_ok=True)

    for file in os.listdir(folder_path):
        if file.endswith('.xlsx'):
            file_path = os.path.join(folder_path, file)
            workbook = load_workbook(filename=file_path, read_only=True)

            data_dict = {}

            for i in range(1, 22):
                sheet_name = workbook.sheetnames[i - 1]
                sheet = workbook[sheet_name]

                for row in sheet.iter_rows(min_row=2, values_only=True):
                    if len(row) < 3:
                        continue

                    key = row[0]
                    value = row[2] if type(row[2]) in [int, float] else 0
                    data_dict[key] = value

            json_file_name = ''.join(filter(str.isdigit, file))[-6:] + '.json'
            json_file_path = os.path.join(json_folder, json_file_name)

            with open(json_file_path, 'w', encoding='utf-8') as json_file:
                json.dump(data_dict, json_file, ensure_ascii=False, indent=4)


folder_path_name = '/Users/alinaaisina/PythonProjects/Test/часть практики/контрольная группа'
process_xlsx_files(folder_path_name)
