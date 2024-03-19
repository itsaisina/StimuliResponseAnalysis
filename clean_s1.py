import os
from openpyxl import load_workbook


def clean_word(word):
    return word.rstrip('.,')


def process_workbook(file_path, output_dir):
    workbook = load_workbook(filename=file_path)
    for sheet in workbook.sheetnames:
        worksheet = workbook[sheet]
        for row in worksheet.iter_rows(min_col=1, max_col=1, min_row=1, values_only=False):
            for cell in row:
                if cell.value:
                    cell.value = clean_word(cell.value)

    output_path = os.path.join(output_dir, os.path.basename(file_path))
    workbook.save(filename=output_path)


def process_all_workbooks(input_dir, output_dir):
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    for file_name in os.listdir(input_dir):
        if file_name.endswith('.xlsx'):
            file_path = os.path.join(input_dir, file_name)
            process_workbook(file_path, output_dir)


file_type_input = int(input('Введите 1 (группа нормы) или 2 (контрольная группа): '))
if file_type_input == 1:
    input_directory = 'группа нормы'
    output_directory = 'folder_группа_нормы'
    process_all_workbooks(input_directory, output_directory)
elif file_type_input == 2:
    input_directory = 'контрольная группа'
    output_directory = 'folder_контрольная_группа'
    process_all_workbooks(input_directory, output_directory)
